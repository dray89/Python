#https://docs.python-guide.org/scenarios/xml/
from __future__ import print_function
import smtplib, openpyxl, sys
from mailmerge import MailMerge
from datetime import date

template = input("Input Document Name and Extension: " )
username = input("Input Sender Email: ")
password = input("Input Password: ")
excel_doc = input("Input Excel Document: ")

from openpyxl import load_workbook
workbook = load_workbook(excel_doc)
worksheet = workbook['Sheet1']
FinalDispositionColumn = worksheet.get_highest_column()
FinalDispositionStatus = worksheet.cell(row=1, column=lastCol).value

   
okay = ["Email Sent", "voicemail left"] #All FinalDispositions Okay to Contact
contacts = {}
for r in range(2, sheet.get_highest_row() + 1):  # Check each contact's best disposition. Check to ensure all Final Dispositions are correct.
    FinalDisposition = sheet.cell(row=r, column=lastCol).value
       if FinalDisposition in okay:
           name = sheet.cell(row=r, column=1).value
           email = sheet.cell(row=r, column=2).value
           contacts[name] = email
        else:
            pass

smtpObj = smtplib.SMTP('smtp.domain', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(username, sys.argv[1])

for name, email in contacts.items():
    from mailmerge import MailMerge
    document = MailMerge(template)
    mergefields = []
        mergefields.append(document.get_merge_fields())
    document.merge(mergefields)

sendmailStatus = smtpObj.sendmail(username, email, body)
if sendmailStatus != {}:
    document('There was a problem sending email to %s: %s' % (email, sendmailStatus))
    
smtpObj.quit()



range = input("How many contacts?: ")
contacts = []
contacts.append(input("Input Contact List: "))
address = []
address.append()


sent_from = "email@domain.com"
to = 
subject = input("Enter Email Subject:" )
body = document(mailmerge)
email_text = """\
From: %s
To: %s
Subject: %s

%s

""" %(sent_from, to, subject, body)  # Log in to email account. Send out emails.
server = smtplib.SMTP_SSL('smtp.domain', 465)
server.ehlo_or_helo_if_needed()
server.login(domain_user, domain_password)
server.sendmail(sent_from, to, email_text)
server.close()

print("Email Sent!")

